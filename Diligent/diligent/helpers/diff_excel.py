"""Excel file diff helper.

Provides structured comparison between two Excel (.xlsx/.xls) files,
returning sheet-level, cell-level, row-level, and named-range differences.

openpyxl is lazy-imported inside function bodies to avoid loading at
module import time (keeps CLI startup under 200ms).
"""

from __future__ import annotations


def diff_excel_summary(path_a: str, path_b: str) -> dict:
    """Compare two Excel files and return structured difference summary.

    Uses openpyxl in read_only + data_only mode for low memory usage.
    Always closes workbooks in finally block.

    Returns dict with keys:
        sheets_changed: int - number of sheets with any differences
        total_sheets: int - total sheets in file B
        cells_differ: int - total cells with different values
        rows_added: int - net rows added across all common sheets
        rows_removed: int - net rows removed across all common sheets
        named_ranges_added: int - named ranges in B not in A
        named_ranges_removed: int - named ranges in A not in B
        changed_sheet_names: list[str] - names of sheets with differences
    """
    from openpyxl import load_workbook

    wb_a = load_workbook(path_a, read_only=True, data_only=True)
    wb_b = load_workbook(path_b, read_only=True, data_only=True)

    try:
        sheets_a = set(wb_a.sheetnames)
        sheets_b = set(wb_b.sheetnames)

        common_sheets = sheets_a & sheets_b
        added_sheets = sheets_b - sheets_a
        removed_sheets = sheets_a - sheets_b

        cells_differ = 0
        rows_added = 0
        rows_removed = 0
        changed_sheet_names: list[str] = []

        # Added/removed sheets count as changed
        changed_sheet_names.extend(sorted(added_sheets))
        changed_sheet_names.extend(sorted(removed_sheets))

        # Compare common sheets
        for sheet_name in sorted(common_sheets):
            ws_a = wb_a[sheet_name]
            ws_b = wb_b[sheet_name]

            sheet_changed = False

            # Read all rows from both sheets
            rows_a = list(ws_a.iter_rows(values_only=True))
            rows_b = list(ws_b.iter_rows(values_only=True))

            min_rows = min(len(rows_a), len(rows_b))

            # Row count differences
            if len(rows_b) > len(rows_a):
                rows_added += len(rows_b) - len(rows_a)
                sheet_changed = True
            elif len(rows_a) > len(rows_b):
                rows_removed += len(rows_a) - len(rows_b)
                sheet_changed = True

            # Compare cell values in overlapping rows
            for i in range(min_rows):
                row_a = rows_a[i]
                row_b = rows_b[i]
                max_cols = max(len(row_a), len(row_b))

                for j in range(max_cols):
                    val_a = row_a[j] if j < len(row_a) else None
                    val_b = row_b[j] if j < len(row_b) else None
                    if val_a != val_b:
                        cells_differ += 1
                        sheet_changed = True

            if sheet_changed:
                changed_sheet_names.append(sheet_name)

        # Named ranges comparison
        names_a = set(wb_a.defined_names.keys())
        names_b = set(wb_b.defined_names.keys())

        named_ranges_added = len(names_b - names_a)
        named_ranges_removed = len(names_a - names_b)

        # Count sheets with changes
        sheets_changed = len(set(changed_sheet_names))

        return {
            "sheets_changed": sheets_changed,
            "total_sheets": len(sheets_b),
            "cells_differ": cells_differ,
            "rows_added": rows_added,
            "rows_removed": rows_removed,
            "named_ranges_added": named_ranges_added,
            "named_ranges_removed": named_ranges_removed,
            "changed_sheet_names": sorted(set(changed_sheet_names)),
        }

    finally:
        wb_a.close()
        wb_b.close()
